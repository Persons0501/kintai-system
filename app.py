import io
import os
from datetime import datetime
from functools import wraps

from flask import (Flask, render_template, request, redirect, url_for,
                   session, flash, send_file)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

import database as db


def calc_break_time_for_excel(total_minutes):
    """Excel合計行用の休憩時間計算"""
    if total_minutes > 480:
        return 60
    elif total_minutes > 360:
        return 45
    return 0


app = Flask(__name__)
app.secret_key = 'kintai-system-secret-key-2026'


# --- デコレータ ---

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if not db.is_user_active(session['user_id']):
            session.clear()
            flash('このアカウントは無効化されています', 'error')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session or not session.get('is_admin'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


# --- ルート ---

@app.route('/')
def index():
    if 'user_id' in session:
        if session.get('is_admin'):
            return redirect(url_for('admin_dashboard'))
        return redirect(url_for('clock'))
    return redirect(url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        login_id = request.form.get('login_id', '').strip()
        password = request.form.get('password', '').strip()

        user = db.authenticate(login_id, password)
        if user:
            session['user_id'] = user['id']
            session['user_name'] = user['name']
            session['is_admin'] = bool(user['is_admin'])
            session['workplace'] = user['workplace']

            if user['is_admin']:
                return redirect(url_for('admin_dashboard'))
            return redirect(url_for('clock'))

        flash('IDまたはパスワードが間違っています', 'error')

    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/clock', methods=['GET', 'POST'])
@login_required
def clock():
    if request.method == 'POST':
        action = request.form.get('action')

        if action == 'clock_in':
            success, msg = db.clock_in(session['user_id'])
            if success:
                flash(f'出勤しました ({msg})', 'success')
            else:
                flash(msg, 'error')

        elif action == 'clock_out':
            success, msg = db.clock_out(session['user_id'])
            if success:
                flash(f'退勤しました ({msg})', 'success')
            else:
                flash(msg, 'error')

        return redirect(url_for('clock'))

    today = db.get_today_status(session['user_id'])
    now = datetime.now()
    return render_template('clock.html', today=today, now=now)


@app.route('/admin')
@admin_required
def admin_dashboard():
    employees = db.get_all_employees()
    workplaces = db.get_all_workplaces()
    today_overview = db.get_today_overview()

    year_month = request.args.get('year_month', datetime.now().strftime('%Y-%m'))
    employee_filter = request.args.get('employee_id', '')
    workplace_filter = request.args.get('workplace', '')

    user_id = int(employee_filter) if employee_filter else None
    records = db.get_attendance_records(user_id=user_id, year_month=year_month)

    if workplace_filter:
        records = [r for r in records if r['workplace'] == workplace_filter]

    return render_template('admin.html',
                           employees=employees,
                           workplaces=workplaces,
                           today_overview=today_overview,
                           records=records,
                           year_month=year_month,
                           employee_filter=employee_filter,
                           workplace_filter=workplace_filter)


@app.route('/admin/add_employee', methods=['POST'])
@admin_required
def add_employee():
    name = request.form.get('name', '').strip()
    workplace = request.form.get('workplace', '').strip()

    if not name:
        flash('名前を入力してください', 'error')
        return redirect(url_for('admin_dashboard'))

    login_id, password = db.create_employee(name, workplace)
    flash(f'従業員を登録しました - 名前: {name} / ID: {login_id} / パスワード: {password}', 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/delete_employee/<int:user_id>', methods=['POST'])
@admin_required
def delete_employee(user_id):
    db.delete_employee(user_id)
    flash('従業員を削除しました', 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/reset_password/<int:user_id>', methods=['POST'])
@admin_required
def reset_password(user_id):
    new_password = db.reset_password(user_id)
    flash(f'パスワードをリセットしました - 新パスワード: {new_password}', 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/edit_employee/<int:user_id>', methods=['POST'])
@admin_required
def edit_employee(user_id):
    name = request.form.get('name', '').strip()
    workplace = request.form.get('workplace', '').strip()
    if not name:
        flash('名前を入力してください', 'error')
        return redirect(url_for('admin_dashboard'))
    db.update_employee(user_id, name, workplace)
    flash(f'{name} の情報を更新しました', 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/edit_attendance', methods=['POST'])
@admin_required
def edit_attendance():
    user_id_str = request.form.get('user_id', '').strip()
    if not user_id_str or not user_id_str.isdigit():
        flash('従業員を選択してください', 'error')
        return redirect(url_for('admin_dashboard'))
    user_id = int(user_id_str)
    date = request.form.get('date', '').strip()
    clock_in = request.form.get('clock_in', '').strip()
    clock_out = request.form.get('clock_out', '').strip()
    if not date:
        flash('日付を入力してください', 'error')
        return redirect(url_for('admin_dashboard'))
    db.update_attendance(user_id, date, clock_in, clock_out)
    flash(f'{date} の勤怠データを更新しました', 'success')
    return redirect(url_for('admin_dashboard'))


@app.route('/admin/download')
@admin_required
def download():
    fmt = request.args.get('format', 'csv')
    year_month = request.args.get('year_month', datetime.now().strftime('%Y-%m'))
    employee_id = request.args.get('employee_id', '')

    workplace = request.args.get('workplace', '')
    user_id = int(employee_id) if employee_id else None
    records = db.get_attendance_records(user_id=user_id, year_month=year_month)

    if workplace:
        records = [r for r in records if r['workplace'] == workplace]

    headers = ['日付', '名前', '勤務先', '出勤', '退勤', '休憩時間', '労働時間']

    if fmt == 'excel':
        wb = Workbook()
        ws = wb.active
        ws.title = '勤怠表'

        # --- スタイル定義 ---
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        title_font = Font(name='Yu Gothic', size=14, bold=True)
        subtitle_font = Font(name='Yu Gothic', size=10, color='666666')
        header_font = Font(name='Yu Gothic', size=10, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='2B3E50', end_color='2B3E50', fill_type='solid')
        cell_font = Font(name='Yu Gothic', size=10)
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        even_fill = PatternFill(start_color='F7F9FC', end_color='F7F9FC', fill_type='solid')
        total_fill = PatternFill(start_color='E8ECF1', end_color='E8ECF1', fill_type='solid')
        total_font = Font(name='Yu Gothic', size=10, bold=True)

        # --- タイトル行 ---
        ws.merge_cells('A1:G1')
        title_cell = ws['A1']
        title_cell.value = f'勤怠表  {year_month}'
        title_cell.font = title_font
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[1].height = 32

        # フィルタ情報
        filter_parts = []
        if workplace:
            filter_parts.append(f'勤務先: {workplace}')
        if user_id:
            emp_name = next((r['name'] for r in records if r['user_id'] == user_id), '')
            if emp_name:
                filter_parts.append(f'従業員: {emp_name}')
        ws.merge_cells('A2:G2')
        sub_cell = ws['A2']
        sub_cell.value = '  '.join(filter_parts) if filter_parts else '全従業員'
        sub_cell.font = subtitle_font
        ws.row_dimensions[2].height = 20

        # 空行
        ws.row_dimensions[3].height = 8

        # --- ヘッダー行 ---
        header_row = 4
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border
        ws.row_dimensions[header_row].height = 28

        # --- データ行 ---
        total_work_minutes = 0
        data_row_count = 0
        for i, r in enumerate(records):
            row_num = header_row + 1 + i
            data_row_count += 1
            row_data = [
                r['date'], r['name'], r['workplace'],
                (r['clock_in'][:5] if r['clock_in'] else ''),
                (r['clock_out'][:5] if r['clock_out'] else ''),
                r['break_time'], r['work_hours']
            ]
            for col_idx, val in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.font = cell_font
                cell.border = thin_border
                if col_idx in (1, 4, 5, 6, 7):
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align
                if i % 2 == 1:
                    cell.fill = even_fill

            total_work_minutes += r.get('total_raw_minutes', 0)

        # --- 合計行 ---
        if data_row_count > 0:
            total_row = header_row + 1 + data_row_count
            total_break = 0
            for r in records:
                if r.get('total_raw_minutes', 0) > 0:
                    total_break += calc_break_time_for_excel(r['total_raw_minutes'])
            net_minutes = total_work_minutes - total_break
            total_h = int(net_minutes // 60)
            total_m = int(net_minutes % 60)

            ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=5)
            total_label = ws.cell(row=total_row, column=1, value='合計')
            total_label.font = total_font
            total_label.fill = total_fill
            total_label.alignment = center_align
            total_label.border = thin_border
            for c in range(2, 6):
                cell = ws.cell(row=total_row, column=c)
                cell.fill = total_fill
                cell.border = thin_border

            break_cell = ws.cell(row=total_row, column=6, value=f'{int(total_break)}分')
            break_cell.font = total_font
            break_cell.fill = total_fill
            break_cell.alignment = center_align
            break_cell.border = thin_border

            work_cell = ws.cell(row=total_row, column=7, value=f'{total_h}時間{total_m}分')
            work_cell.font = total_font
            work_cell.fill = total_fill
            work_cell.alignment = center_align
            work_cell.border = thin_border

            ws.row_dimensions[total_row].height = 28

        # --- 列幅 ---
        col_widths = {'A': 14, 'B': 18, 'C': 16, 'D': 10, 'E': 10, 'F': 12, 'G': 16}
        for col_letter, width in col_widths.items():
            ws.column_dimensions[col_letter].width = width

        # --- 印刷設定 ---
        ws.sheet_properties.pageSetUpPr = None
        ws.page_setup.orientation = 'landscape'
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'勤怠表_{year_month}.xlsx'
        return send_file(output, as_attachment=True, download_name=filename,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    else:  # CSV
        import csv
        output = io.StringIO()
        # BOM付きUTF-8（Excelで文字化けしないように）
        output.write('\ufeff')
        writer = csv.writer(output)
        writer.writerow(headers)

        for r in records:
            writer.writerow([
                r['date'], r['name'], r['workplace'],
                (r['clock_in'][:5] if r['clock_in'] else ''),
                (r['clock_out'][:5] if r['clock_out'] else ''),
                r['break_time'], r['work_hours']
            ])

        output.seek(0)
        filename = f'kintai_{year_month}.csv'
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8')),
            as_attachment=True,
            download_name=filename,
            mimetype='text/csv; charset=utf-8'
        )


if __name__ == '__main__':
    db.init_db()
    print('=== 勤怠管理システム ===')
    print('http://localhost:5000 でアクセスしてください')
    print('管理者ログイン → ID: admin / パスワード: admin123')
    print('========================')

    from waitress import serve
    serve(app, host='0.0.0.0', port=5000, threads=8)
